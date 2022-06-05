from curses.ascii import US
import os
from readline import read_init_file
from secrets import choice
import shutil
from openpyxl import load_workbook


def load_book():
    path_test = os.path.join(os.path.expanduser('~'), 'Projects', 'mangatracker', 'files', 'Manga Collection.xlsx')
    workBook = load_workbook(path_test, read_only=False)
    return path_test, workBook

def get_titles(user_input):
    if user_input == 'a':
        col = 26                    ## column 26=Y, AC link
        print("You chose AC")
    elif user_input == 'm':
        col = 25                    ## column 25=Y, MAL link
        print("You chose MAL")
    path_test, book = load_book()
    sheet = book['auto']
    mangalist = []
    for row in sheet.iter_rows(min_row=2, min_col=col): 
        mangalist.append(row[0].value)
    return mangalist

def copy_to_cloud(src, dst):
    choice = None
    while choice not in {"s","n"}:
        choice = input("Copiare? (s/n): ")
    
    if choice  == "s":
        shutil.copy2(src, dst)
        print("File copiato nel cloud.")
    elif choice  == "n":
        print("File non copiato.")
        