import os
from datetime import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook

def get_titles():
    print("Reading titles list from excel file...")
    path = os.path.join(os.path.expanduser('~'), 'coding', 'files', 'Manga Collection 2021.xlsx')
    workBook = load_workbook(path, read_only=False)
    workSheet = workBook['auto']
    mangaLinks = []
    for row in workSheet.iter_rows(min_row=2, min_col=25):
        mangaLinks.append(row[0].value)
    return mangaLinks

def get_metadata():
    print("Initializing webdriver...")
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--headless')
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument('--disable-dev-shm-usage')
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    # Apro i link dalla lista e recupero ratings, members, rank, popularity
    print("Fetching data...")
    ratings = []
    members = []
    rankings = []
    popularities = []

    for item_link in get_titles():
        print(item_link)
        driver.get(item_link)
        wait = WebDriverWait(driver, 5)
        rating = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[class*="score-label"]')))
        member = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'span[class="numbers members"] strong')))
        ranking = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'span[class="numbers ranked"] strong')))
        popularity = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'span[class="numbers popularity"] strong')))
        ratings.append(rating.text)
        members.append(member.text)
        rankings.append(ranking.text)
        popularities.append(popularity.text)
    # return ratings, members, rankings, popularities


    print("Writing data to excel file...")
    path = os.path.join(os.path.expanduser('~'), 'coding', 'files', 'Manga Collection 2021.xlsx')
    workBook = load_workbook(path, read_only=False)
    workSheet = workBook['Lista']
    now = datetime.now()
    timestamp = now.strftime("%d/%m/%Y %H:%M:%S")

    workSheet.cell(row=1, column=20, value="ratings_mal")       # colonna T
    for row_num, value in enumerate(ratings, start=2):
        workSheet.cell(row=row_num, column=20, value=value)

    workSheet.cell(row=1, column=21, value="members_MAL")       # colonna U
    for row_num, value in enumerate(members, start=2):
        workSheet.cell(row=row_num, column=21, value=value)

    workSheet.cell(row=1, column=22, value="ranking_MAL")       # colonna V
    for row_num, value in enumerate(rankings, start=2):
        workSheet.cell(row=row_num, column=22, value=value)

    workSheet.cell(row=1, column=23, value="popularity_MAL")    # colonna W
    for row_num, value in enumerate(popularities, start=2):
        workSheet.cell(row=row_num, column=23, value=value)

    workSheet.cell(row=1, column=25, value=timestamp)               # colonna Y

    workBook.save(path)


    print("Stopping webdriver...")
    driver.stop_client()
    driver.close()
    driver.quit()
    print("Program stopped.")

