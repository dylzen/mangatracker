import os
from time import sleep
import config
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import requests
import re
from bs4 import BeautifulSoup

def get_manga():
    print("Initializing webdriver...")
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-dev-shm-usage')
    driver = webdriver.Chrome(options=chrome_options)

    print("Reading titles list from excel file...")
    path_test = os.path.join(os.path.expanduser('~'), 'coding', 'files', 'Manga Collection 2021.xlsx')
    workBook = load_workbook(path_test, read_only=False)
    workSheet = workBook['auto']

    ### Carico lista dei links dalla colonna 26=Z
    mangalist = []
    for row in workSheet.iter_rows(min_row=2, min_col=26):
        mangalist.append(row[0].value)

    print("Fetching data...")
    home_url = config.ac_home_url
    titles_ita = []
    stories = []
    drawings = []
    categories = []
    years = []
    volumes = []
    italy_statuses = []
    next_releases = []
    next_releases_long = []
    next_releases_dates = []
    latest_releases = []
    latest_releases_dates = []

    for manga in mangalist:
        print("Fetching "+manga)
        response = requests.get(manga, allow_redirects=True)
        soup = BeautifulSoup(response.text, 'html.parser')
        if soup.find(text="Informazione Pubblicitaria - "):
            ### This is to bypass the video advert that sometimes appears when loading the url ###
            session = requests.Session()
            response = session.get(manga)
            sleep(1)
            response = session.get(manga)
            soup = BeautifulSoup(response.text, 'html.parser')
        titolo_italiano = soup.find('h1').getText()
        print(titolo_italiano)
        storia = soup.find(text="Storia")
        disegni = soup.find(text="Disegni")
        categoria = soup.find(text="Categoria")
        anno = soup.find(text="Anno")
        volumi = soup.find(text=re.compile("Volumi"))
        stato_ita = soup.find(text="Stato in Italia")
        
        storia_td = storia.parent
        disegni_td = disegni.parent
        categoria_td = categoria.parent
        anno_td = anno.parent
        volumi_td = volumi.parent
        stato_ita_td = stato_ita.parent

        target_sto = storia_td.find_next_sibling('dd').text
        target_dis = disegni_td.find_next_sibling('dd').text
        target_cat = categoria_td.find_next_sibling('dd').text
        target_anno = anno_td.find_next_sibling('dd').text
        target_volumi = volumi_td.find_next_sibling('dd').text
        target_statoIt = stato_ita_td.find_next_sibling('dd').text

        titles_ita.append(titolo_italiano.strip())
        stories.append(target_sto.strip())
        drawings.append(target_dis.strip())
        target_cat = re.sub("\s+", " ", target_cat.strip())
        categories.append(target_cat.strip())
        target_anno = re.sub("\s+", "", target_anno.strip())
        years.append(target_anno.strip())
        volumes.append(target_volumi.strip())
        italy_statuses.append(target_statoIt.strip())

        if ((target_statoIt.strip() == "in corso") or (target_statoIt.strip() == "annunciato")) and soup.find(text="Prossima uscita") is not None:
            prossima_uscita = soup.find('h3').getText()
            next_releases_long.append(prossima_uscita.strip())
            next_release_link = soup.select_one("a[href*=edizione\/]")
            next_releases_dates.append(home_url+next_release_link.get('href'))
            latest_releases.append("")
            latest_releases_dates.append("N.D.")
        elif target_statoIt.strip() == "completato" and soup.find(text="Prossima uscita") is not None:
            prossima_uscita = soup.find('h3').getText()
            next_releases_long.append(prossima_uscita.strip())
            next_release_link = soup.select_one("a[href*=edizione\/]")
            next_releases_dates.append(home_url+next_release_link.get('href'))
            latest_releases.append("")
            latest_releases_dates.append("N.D.")
        elif soup.find(text="Ultima uscita") is not None:
            ultima_uscita = soup.find('h3').getText()
            latest_releases.append("Ultima uscita: "+ultima_uscita.strip())
            next_release_link = soup.select_one("a[href*=edizione\/]")
            latest_releases_dates.append(home_url+next_release_link.get('href'))
            next_releases.append("")
            next_releases_long.append("")
            next_releases_dates.append("N.D.")
        else:
            next_releases.append("")
            next_releases_long.append("")
            next_releases_dates.append("N.D.")
            latest_releases.append("")
            latest_releases_dates.append("N.D.")
            continue

    print("Creating dates list...")
    next_volume_dates = []
    latest_volume_dates = []
    session = requests.Session()
    for item in next_releases_dates:
        if item != "N.D.":
            response_next = session.get(item)
            soup_next = BeautifulSoup(response_next.text, 'html.parser')
            next_date_parent = soup_next.find('strong', text="Data pubblicazione:")
            next_volume_dates.append(next_date_parent.next_sibling.text)
        else:
            next_volume_dates.append("")

    for item in latest_releases_dates:
        if item != "N.D.":
            response_latest = session.get(item)
            soup_latest = BeautifulSoup(response_latest.text, 'html.parser')
            latest_date_parent = soup_latest.find('strong', text="Data pubblicazione:")
            latest_volume_dates.append(latest_date_parent.next_sibling.text)
        else:
            latest_volume_dates.append("")

    dates_new = []
    for date in next_volume_dates:
        dates_new.append(date.replace('/01/', 'gen').replace('/11/', 'nov').replace('/12/', 'dic'))

    print("Writing data to excel file...")
    book = load_workbook(path_test)
    sheet = book['auto']
    column = 1
    sheet.cell(row=1, column=column, value="Titolo italiano")
    for i, value in enumerate(titles_ita, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 2
    sheet.cell(row=1, column=column, value="Storia")
    for i, value in enumerate(stories, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 3
    sheet.cell(row=1, column=column, value="Disegni")
    for i, value in enumerate(drawings, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 4
    sheet.cell(row=1, column=column, value="Categoria")
    for i, value in enumerate(categories, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 5
    sheet.cell(row=1, column=column, value="Anno")
    for i, value in enumerate(years, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 6
    sheet.cell(row=1, column=column, value="Volumi pubblicati")
    for i, value in enumerate(volumes, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 7
    sheet.cell(row=1, column=column, value="Ultimo volume")
    for i, value in enumerate(latest_releases, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 8
    sheet.cell(row=1, column=column, value="Ultima data di uscita")
    for i, value in enumerate(latest_volume_dates, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 9
    sheet.cell(row=1, column=column, value="Prossimo volume")
    for i, value in enumerate(next_releases_long, start=1):
        sheet.cell(row=i+1, column=column, value=value)
    column = 10
    sheet.cell(row=1, column=column, value="Prossima data di uscita")
    for i, value in enumerate(next_volume_dates, start=1):
        sheet.cell(row=i+1, column=column, value=value)

    book.save(path_test)

    print("Stopping webdriver...")
    driver.stop_client()
    driver.close()
    driver.quit()

    print("Program stopped.")